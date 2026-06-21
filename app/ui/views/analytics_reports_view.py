# app/ui/views/analytics_reports_view.py
import math
from datetime import datetime, time
from decimal import Decimal

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QFrame, QComboBox, QDateEdit, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QMessageBox, QFileDialog, QGraphicsDropShadowEffect,
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor, QTextDocument
from PySide6.QtPrintSupport import QPrinter

from app.models.schema import Sale, Product, User

# Pagination configuration
PAGE_SIZE = 50

def format_currency(value):
    """KES amount with thousands separators, e.g. KES 12,345.00"""
    return f"KES {float(value):,.2f}"

def _apply_card_shadow(widget):
    """Soft drop shadow shared with the dashboard's card styling."""
    shadow = QGraphicsDropShadowEffect(widget)
    shadow.setBlurRadius(20)
    shadow.setOffset(0, 6)
    shadow.setColor(QColor(31, 36, 48, 28))
    widget.setGraphicsEffect(shadow)


class AnalyticsReportsView(QWidget):
    """
    Filterable sales report — by product, cashier, and date range — with
    on-demand PDF and Excel export. Supports pagination.
    """

    def __init__(self, user, db_session):
        super().__init__()
        self.user = user
        self.db = db_session
        self._current_page = 1
        self._total_pages = 1
        self._current_rows = []
        
        self.init_ui()
        self.run_report()

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def init_ui(self):
        self.setStyleSheet("""
            QWidget {
                background-color: #f4f6f8; font-family: 'Segoe UI', Arial, sans-serif;
                color: #1f2430; font-size: 13px;
            }
            QFrame#Card { background-color: #ffffff; border-radius: 14px; border: 1px solid #eef0f3; }
            QLabel#CardTitle { font-size: 14px; font-weight: 700; color: #1f2430; }
            QLabel#FieldLabel { font-size: 11px; color: #8a8f98; font-weight: 600; }
            QComboBox, QDateEdit {
                border: 1px solid #d8dce2; padding: 7px 10px; border-radius: 8px; background: white;
            }
            QComboBox:focus, QDateEdit:focus { border: 1.5px solid #00adb5; }
            QTableWidget {
                background-color: #ffffff; border: none; gridline-color: transparent;
                alternate-background-color: #fafbfc; selection-background-color: #e6f9fa;
            }
            QHeaderView::section {
                background-color: #fafbfc; color: #6b7280; font-weight: 700; font-size: 11px;
                padding: 10px 8px; border: none; border-bottom: 1px solid #e7eaee;
            }
            QTableWidget::item { padding: 7px 4px; border-bottom: 1px solid #f0f2f5; }
            QPushButton {
                border-radius: 8px; font-weight: 600; padding: 9px 16px; border: none;
                background: #eef1f4; color: #1f2430;
            }
            QPushButton:hover { background: #e3e7ec; }
            QPushButton#PrimaryButton { background: #00adb5; color: white; }
            QPushButton#PrimaryButton:hover { background: #019aa1; }
            QPushButton#PdfButton { background: #fde2e1; color: #c0392b; }
            QPushButton#PdfButton:hover { background: #fbd2d0; }
            QPushButton#ExcelButton { background: #e3f6e8; color: #2a9d6f; }
            QPushButton#ExcelButton:hover { background: #d4f0db; }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(18)

        header_row = QHBoxLayout()
        title_lbl = QLabel("Sales Report")
        title_lbl.setStyleSheet("font-size: 22px; font-weight: 800; color: #1f2430;")
        header_row.addWidget(title_lbl)
        header_row.addStretch()
        layout.addLayout(header_row)

        layout.addWidget(self._build_filter_card())
        layout.addWidget(self._build_summary_card())
        layout.addWidget(self._build_results_card(), stretch=1)

    def _build_filter_card(self):
        card = QFrame()
        card.setObjectName("Card")
        outer = QVBoxLayout(card)
        outer.setContentsMargins(20, 18, 20, 18)
        outer.setSpacing(12)

        title = QLabel("Filters")
        title.setObjectName("CardTitle")
        outer.addWidget(title)

        filter_row = QHBoxLayout()
        filter_row.setSpacing(16)

        # Inputs
        self.product_combo = QComboBox()
        self.product_combo.setEditable(True)
        self.product_combo.setInsertPolicy(QComboBox.NoInsert)
        self.user_combo = QComboBox()
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addDays(-30))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())

        for lbl, widget in [("PRODUCT", self.product_combo), ("CASHIER", self.user_combo), ("FROM", self.date_from), ("TO", self.date_to)]:
            box = QVBoxLayout()
            l = QLabel(lbl)
            l.setObjectName("FieldLabel")
            box.addWidget(l)
            box.addWidget(widget)
            filter_row.addLayout(box, 1 if lbl != "PRODUCT" else 2)
        outer.addLayout(filter_row)

        action_row = QHBoxLayout()
        run_btn = QPushButton("Run Report")
        run_btn.setObjectName("PrimaryButton")
        run_btn.clicked.connect(self._reset_and_run)
        
        reset_btn = QPushButton("Reset Filters")
        reset_btn.clicked.connect(self.reset_filters)

        action_row.addWidget(run_btn)
        action_row.addWidget(reset_btn)
        action_row.addStretch()
        action_row.addWidget(QPushButton("Export PDF", objectName="PdfButton", clicked=self.export_pdf))
        action_row.addWidget(QPushButton("Export Excel", objectName="ExcelButton", clicked=self.export_excel))
        outer.addLayout(action_row)

        self._populate_filter_options()
        _apply_card_shadow(card)
        return card

    def _build_summary_card(self):
        card = QFrame()
        card.setObjectName("SummaryCard")
        card.setStyleSheet("QFrame#SummaryCard { background-color: #f0fbfb; border-radius: 14px; border: 1px solid #c7eef0; }")
        layout = QHBoxLayout(card)
        layout.setContentsMargins(20, 14, 20, 14)
        self.summary_lbl = QLabel("0 transactions • 0 units • KES 0.00")
        self.summary_lbl.setStyleSheet("font-size: 14px; font-weight: 700; color: #017a80;")
        layout.addWidget(self.summary_lbl)
        _apply_card_shadow(card)
        return card

    def _build_results_card(self):
        card = QFrame()
        card.setObjectName("Card")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 18, 20, 18)
        
        self.results_table = QTableWidget(0, 6)
        self.results_table.setHorizontalHeaderLabels(["Date", "Receipt #", "Product", "Cashier", "Qty", "Subtotal"])
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.results_table.verticalHeader().setVisible(False)
        self.results_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        layout.addWidget(self.results_table)

        # Pagination Bar
        pager_layout = QHBoxLayout()
        self.prev_btn = QPushButton("‹ Prev")
        self.prev_btn.setFixedWidth(80)
        self.prev_btn.clicked.connect(lambda: self._go_page(-1))
        self.page_lbl = QLabel("Page 1 of 1")
        self.next_btn = QPushButton("Next ›")
        self.next_btn.setFixedWidth(80)
        self.next_btn.clicked.connect(lambda: self._go_page(1))
        pager_layout.addStretch()
        pager_layout.addWidget(self.prev_btn)
        pager_layout.addWidget(self.page_lbl)
        pager_layout.addWidget(self.next_btn)
        pager_layout.addStretch()
        layout.addLayout(pager_layout)

        _apply_card_shadow(card)
        return card

    # ------------------------------------------------------------------
    # Data Logic
    # ------------------------------------------------------------------

    def _populate_filter_options(self):
        self.product_combo.addItem("All Products", None)
        for p in self.db.query(Product).order_by(Product.name.asc()).all():
            self.product_combo.addItem(p.name, p.id)
        self.user_combo.addItem("All Cashiers", None)
        for u in self.db.query(User).order_by(User.name.asc()).all():
            self.user_combo.addItem(u.name, u.id)

    def _reset_and_run(self):
        self._current_page = 1
        self.run_report()

    def reset_filters(self):
        self.product_combo.setCurrentIndex(0)
        self.user_combo.setCurrentIndex(0)
        self.date_from.setDate(QDate.currentDate().addDays(-30))
        self.date_to.setDate(QDate.currentDate())
        self._current_page = 1
        self.run_report()

    def _go_page(self, delta):
        self._current_page += delta
        self.run_report()

    def _get_base_query(self):
        range_start = datetime.combine(self.date_from.date().toPython(), time.min)
        range_end = datetime.combine(self.date_to.date().toPython(), time.max)
        query = self.db.query(Sale).filter(Sale.sale_date >= range_start, Sale.sale_date <= range_end)
        if self.product_combo.currentData() is not None:
            query = query.filter(Sale.product_id == self.product_combo.currentData())
        if self.user_combo.currentData() is not None:
            query = query.filter(Sale.user_id == self.user_combo.currentData())
        return query

    def run_report(self):
        query = self._get_base_query()
        
        # Paginate
        total_count = query.count()
        self._total_pages = max(1, math.ceil(total_count / PAGE_SIZE))
        self._current_page = max(1, min(self._current_page, self._total_pages))
        
        # Get current page
        self._current_rows = query.order_by(Sale.sale_date.desc()).offset((self._current_page - 1) * PAGE_SIZE).limit(PAGE_SIZE).all()
        
        # Get full for summary
        full_results = query.all()
        self._render_results(full_results)

    def _render_results(self, full_results):
        self.results_table.setRowCount(0)
        self.page_lbl.setText(f"Page {self._current_page} of {self._total_pages}")
        self.prev_btn.setEnabled(self._current_page > 1)
        self.next_btn.setEnabled(self._current_page < self._total_pages)

        count = len(full_results)
        total_qty = sum(s.quantity for s in full_results)
        total_revenue = sum(Decimal(str(s.quantity)) * Decimal(str(s.price)) for s in full_results)
        self.summary_lbl.setText(f"{count:,} transactions • {total_qty:,} units • {format_currency(total_revenue)} total")

        self.results_table.setRowCount(len(self._current_rows))
        for idx, sale in enumerate(self._current_rows):
            subtotal = Decimal(str(sale.quantity)) * Decimal(str(sale.price))
            self.results_table.setItem(idx, 0, QTableWidgetItem(sale.sale_date.strftime("%Y-%m-%d %H:%M")))
            self.results_table.setItem(idx, 1, QTableWidgetItem(sale.receipt.receipt_number if sale.receipt else "—"))
            self.results_table.setItem(idx, 2, QTableWidgetItem(sale.product.name if sale.product else "Deleted"))
            self.results_table.setItem(idx, 3, QTableWidgetItem(sale.user.name if sale.user else "System"))
            self.results_table.setItem(idx, 4, QTableWidgetItem(f"{sale.quantity:,}"))
            self.results_table.setItem(idx, 5, QTableWidgetItem(format_currency(subtotal)))

    # ------------------------------------------------------------------
    # Exports
    # ------------------------------------------------------------------

    def export_pdf(self):
        rows = self._get_base_query().all()
        if not rows: return QMessageBox.information(self, "Empty", "No data to export.")
        
        path, _ = QFileDialog.getSaveFileName(self, "Export PDF", "sales_report.pdf", "PDF Files (*.pdf)")
        if path:
            doc = QTextDocument()
            doc.setHtml(self._build_report_html(rows))
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(path)
            doc.print_(printer)
            QMessageBox.information(self, "Success", "Exported successfully.")

    def export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except: return QMessageBox.critical(self, "Error", "Install openpyxl")
        
        rows = self._get_base_query().all()
        if not rows: return QMessageBox.information(self, "Empty", "No data to export.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        headers = ["Date", "Receipt", "Product", "Cashier", "Qty", "Price", "Subtotal"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="00ADB5", end_color="00ADB5", fill_type="solid")
        for i, cell in enumerate(ws[1], 1):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        total_qty = 0
        total_revenue = Decimal("0.00")
        for s in rows:
            subtotal = Decimal(str(s.quantity)) * Decimal(str(s.price))
            total_qty += s.quantity
            total_revenue += subtotal
            ws.append([s.sale_date.strftime("%Y-%m-%d %H:%M"), s.receipt.receipt_number if s.receipt else "-", 
                       s.product.name if s.product else "-", s.user.name if s.user else "-", 
                       s.quantity, float(s.price), float(subtotal)])
        
        ws.append(["", "", "", "TOTALS", total_qty, "", float(total_revenue)])
        wb.save(path := QFileDialog.getSaveFileName(self, "Export Excel", "sales_report.xlsx", "Excel Files (*.xlsx)")[0])
        if path: QMessageBox.information(self, "Success", "Exported successfully.")

    def _build_report_html(self, rows):
        rows_html = "".join([f"""
            <tr>
                <td>{s.sale_date.strftime('%Y-%m-%d %H:%M')}</td>
                <td>{s.receipt.receipt_number if s.receipt else '-'}</td>
                <td>{s.product.name if s.product else '-'}</td>
                <td>{s.user.name if s.user else '-'}</td>
                <td style="text-align:right;">{s.quantity:,}</td>
                <td style="text-align:right;">{format_currency(Decimal(str(s.quantity)) * Decimal(str(s.price)))}</td>
            </tr>""" for s in rows])
        
        total_qty = sum(s.quantity for s in rows)
        total_rev = sum(Decimal(str(s.quantity)) * Decimal(str(s.price)) for s in rows)
        
        return f"""<html><head><style>
            table {{ width: 100%; border-collapse: collapse; }}
            th {{ background: #00adb5; color: white; padding: 8px; }}
            td {{ border-bottom: 1px solid #ddd; padding: 6px; }}
            tfoot {{ font-weight: bold; }}
        </style></head><body><h1>Sales Report</h1>
        <table><thead><tr><th>Date</th><th>Receipt</th><th>Product</th><th>Cashier</th><th>Qty</th><th>Subtotal</th></tr></thead>
        <tbody>{rows_html}</tbody>
        <tfoot><tr><td colspan="4">TOTALS</td><td style="text-align:right;">{total_qty:,}</td><td style="text-align:right;">{format_currency(total_rev)}</td></tr></tfoot>
        </table></body></html>"""